import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_evaluation_spreadsheet(filename):
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Colors
    header_fill = PatternFill(start_color="111420", end_color="111420", fill_type="solid")
    primary_fill = PatternFill(start_color="00F59B", end_color="00F59B", fill_type="solid")
    accent_fill = PatternFill(start_color="1E2235", end_color="1E2235", fill_type="solid")
    stripe_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_header_dark = Font(name="Calibri", size=11, bold=True, color="0B130E")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_title = Font(name="Calibri", size=16, bold=True, color="00A86B")
    font_subtitle = Font(name="Calibri", size=11, color="64748B")
    font_regular = Font(name="Calibri", size=11)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Players list from SQLite database
    players = [
        {"id": 1, "name": "Thiago Silva", "nick": "Fela", "pos": "ATA"},
        {"id": 2, "name": "Elias Lemes", "nick": "Elias", "pos": "MEI"},
        {"id": 3, "name": "Jhonatan", "nick": "Jho", "pos": "MEI"},
        {"id": 4, "name": "Lucas Hagen", "nick": "Hagen", "pos": "MEI"},
        {"id": 5, "name": "William Stocco", "nick": "Stocco", "pos": "ZAG"},
        {"id": 6, "name": "Wesley", "nick": "Wesley", "pos": "ATA"},
        {"id": 7, "name": "Pedro", "nick": "Pedro", "pos": "VOL"},
        {"id": 8, "name": "Rafael", "nick": "Rafão", "pos": "ZAG"},
        {"id": 9, "name": "Wellington", "nick": "Wellington", "pos": "MEI"},
        {"id": 10, "name": "Willian Calebe", "nick": "Calebe", "pos": "ATA"},
        {"id": 11, "name": "Yan", "nick": "Yan", "pos": "MEI"},
        {"id": 12, "name": "Yuri", "nick": "Yuri", "pos": "LAT"},
        {"id": 13, "name": "Adelarme", "nick": "Adelarme", "pos": "ZAG"},
        {"id": 14, "name": "Eryc", "nick": "Eryc", "pos": "VOL"},
        {"id": 15, "name": "Babycholo", "nick": "Babycholo", "pos": "MEI"},
        {"id": 16, "name": "BND", "nick": "BND", "pos": "ATA"},
        {"id": 17, "name": "Guilherme Krolow", "nick": "Krolow", "pos": "MEI"},
        {"id": 18, "name": "Hilário", "nick": "Hilário", "pos": "GOL"},
        {"id": 19, "name": "Icarus", "nick": "Icarus", "pos": "MEI"}
    ]

    # ==========================================
    # SHEET 1: INSTRUÇÕES
    # ==========================================
    ws1 = wb.create_sheet(title="Como Avaliar")
    ws1.views.sheetView[0].showGridLines = True

    ws1["B2"] = "⚽ plugshawtycafetoes FC — Avaliação Oficial de Atributos"
    ws1["B2"].font = Font(name="Calibri", size=18, bold=True, color="008040")
    ws1["B3"] = "Sistema de Votação Anônima para Geração das Cartas FUT e Balanceador de Equipes"
    ws1["B3"].font = font_subtitle

    instructions = [
        ("COMO FUNCIONA O SISTEMA:", Font(name="Calibri", size=12, bold=True)),
        ("1. O voto é 100% ANÔNIMO. Ninguém saberá quem deu qual nota para quem.", font_regular),
        ("2. Cada avaliador dá notas de 1 a 10 para cada um dos 6 atributos dos 19 atletas.", font_regular),
        ("3. As notas de 1 a 10 são multiplicadas por 10 e convertidas automaticamente para as Cartas FUT (0 a 99).", font_regular),
        ("4. O OVR Final é a média equilibrada dos 6 fundamentos de cada jogador.", font_regular),
        ("", font_regular),
        ("OS 6 ATRIBUTOS DA CARTA FUT:", Font(name="Calibri", size=12, bold=True)),
        ("🏃 PAC (Ritmo / Velocidade):", font_bold),
        ("   Arranque, aceleração nos primeiros metros e fôlego para correr o jogo todo.", font_regular),
        ("⚽ SHO (Finalização / Chute):", font_bold),
        ("   Pontaria, força no chute de curta e longa distância, facilidade para marcar gols.", font_regular),
        ("👟 PAS (Passe / Visão de Jogo):", font_bold),
        ("   Precisão no passe curto, inversões de bola, cruzamentos e visão para assistência.", font_regular),
        ("🪄 DRI (Drible / Agilidade):", font_bold),
        ("   Controle de bola sob pressão, domínio limpo, facilidade no 1 contra 1 e ginga.", font_regular),
        ("🛡️ DEF (Defesa / Marcação):", font_bold),
        ("   Desarme pontual, interceptação, combate físico na marcação e senso de cobertura.", font_regular),
        ("💪 PHY (Físico / Raça):", font_bold),
        ("   Força no ombro a ombro, vigor em disputas, raça e resistência física aos choques.", font_regular),
        ("", font_regular),
        ("CRITÉRIO DE NOTAS (ESCALA 1 A 10):", Font(name="Calibri", size=12, bold=True)),
        ("• 1 a 4: Ponto fraco do jogador / quase não utiliza", font_regular),
        ("• 5 a 6: Mediano / padrão comum do futebol", font_regular),
        ("• 7 a 8: Bom fundamento / jogador se destaca nesse aspecto", font_regular),
        ("• 9 a 10: Craque incontestável do grupo nesse fundamento", font_regular),
    ]

    r = 5
    for text, f in instructions:
        ws1.cell(row=r, column=2, value=text).font = f
        r += 1

    ws1.column_dimensions['A'].width = 4
    ws1.column_dimensions['B'].width = 85

    # ==========================================
    # SHEET 2: VOTAÇÃO ANÔNIMA
    # ==========================================
    ws2 = wb.create_sheet(title="Votacao")
    ws2.views.sheetView[0].showGridLines = True

    ws2["A1"] = "📝 REGISTRO DE VOTOS ANÔNIMOS"
    ws2["A1"].font = font_title
    ws2["A2"] = "Preencha suas notas de 1 a 10 para cada atleta. O campo 'Votante' pode ser 'Anon 1', 'Anon 2' ou em branco!"
    ws2["A2"].font = font_subtitle

    v_headers = [
        "Avaliador (Anônimo)", "Jogador Avaliado", "PAC (Velocidade)", 
        "SHO (Finalização)", "PAS (Passe)", "DRI (Drible)", 
        "DEF (Marcação)", "PHY (Físico)", "Média do Voto"
    ]

    for col_num, h_title in enumerate(v_headers, 1):
        cell = ws2.cell(row=4, column=col_num, value=h_title)
        cell.fill = header_fill
        cell.font = font_header
        cell.alignment = align_center

    ws2.row_dimensions[4].height = 26

    # Pre-generate structured template for up to 10 anonymous voters x 19 players
    curr_row = 5
    for voter_idx in range(1, 11):
        voter_label = f"Votante {voter_idx:02d}"
        for p in players:
            ws2.cell(row=curr_row, column=1, value=voter_label).alignment = align_center
            ws2.cell(row=curr_row, column=2, value=p["name"]).alignment = align_left
            
            for c in range(3, 9):
                cell = ws2.cell(row=curr_row, column=c)
                cell.alignment = align_center
            
            # Formula for vote average: =IFERROR(AVERAGE(C5:H5), "")
            avg_cell = ws2.cell(row=curr_row, column=9)
            avg_cell.value = f'=IFERROR(AVERAGE(C{curr_row}:H{curr_row}), "")'
            avg_cell.alignment = align_center
            avg_cell.number_format = '0.0'
            avg_cell.font = font_bold
            
            # Alternating row colors
            fill = stripe_fill if (curr_row % 2 == 0) else white_fill
            for c in range(1, 10):
                ws2.cell(row=curr_row, column=c).border = thin_border
                if c < 9:
                    ws2.cell(row=curr_row, column=c).fill = fill

            curr_row += 1

    # ==========================================
    # SHEET 3: CONSOLIDADO & CARTAS FUT
    # ==========================================
    ws3 = wb.create_sheet(title="Cartas FUT (Consolidado)")
    ws3.views.sheetView[0].showGridLines = True

    ws3["A1"] = "🎴 CARTAS FUT DO ELENCO — PLUGSHAWTYCAFETOES FC"
    ws3["A1"].font = font_title
    ws3["A2"] = "Médias calculadas automaticamente a partir da aba 'Votacao'. OVR Oficial de cada atleta."
    ws3["A2"].font = font_subtitle

    c_headers = [
        "ID", "Nome Oficial", "Apelido", "Posição", "Nº Votos",
        "PAC (Ritmo)", "SHO (Chute)", "PAS (Passe)", "DRI (Drible)", 
        "DEF (Defesa)", "PHY (Físico)", "OVR FINAL", "Tier da Carta"
    ]

    for col_num, h_title in enumerate(c_headers, 1):
        cell = ws3.cell(row=4, column=col_num, value=h_title)
        if h_title == "OVR FINAL":
            cell.fill = primary_fill
            cell.font = font_header_dark
        elif h_title in ["PAC (Ritmo)", "SHO (Chute)", "PAS (Passe)", "DRI (Drible)", "DEF (Defesa)", "PHY (Físico)"]:
            cell.fill = accent_fill
            cell.font = font_header
        else:
            cell.fill = header_fill
            cell.font = font_header
        cell.alignment = align_center

    ws3.row_dimensions[4].height = 28

    for idx, p in enumerate(players, 5):
        ws3.cell(row=idx, column=1, value=p["id"]).alignment = align_center
        ws3.cell(row=idx, column=2, value=p["name"]).alignment = align_left
        ws3.cell(row=idx, column=3, value=p["nick"]).alignment = align_center
        ws3.cell(row=idx, column=4, value=p["pos"]).alignment = align_center

        # Count votes formula
        ws3.cell(row=idx, column=5, value=f'=COUNTIFS(Votacao!B:B, B{idx}, Votacao!C:C, ">0")').alignment = align_center

        # PAC: =IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!C:C)*10, 0), 50)
        ws3.cell(row=idx, column=6, value=f'=IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!C:C)*10, 0), 50)').alignment = align_center
        # SHO: =IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!D:D)*10, 0), 50)
        ws3.cell(row=idx, column=7, value=f'=IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!D:D)*10, 0), 50)').alignment = align_center
        # PAS: =IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!E:E)*10, 0), 50)
        ws3.cell(row=idx, column=8, value=f'=IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!E:E)*10, 0), 50)').alignment = align_center
        # DRI: =IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!F:F)*10, 0), 50)
        ws3.cell(row=idx, column=9, value=f'=IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!F:F)*10, 0), 50)').alignment = align_center
        # DEF: =IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!G:G)*10, 0), 50)
        ws3.cell(row=idx, column=10, value=f'=IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!G:G)*10, 0), 50)').alignment = align_center
        # PHY: =IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!H:H)*10, 0), 50)
        ws3.cell(row=idx, column=11, value=f'=IFERROR(ROUND(AVERAGEIF(Votacao!B:B, B{idx}, Votacao!H:H)*10, 0), 50)').alignment = align_center

        # OVR: =ROUND(AVERAGE(F{idx}:K{idx}), 0)
        ovr_cell = ws3.cell(row=idx, column=12, value=f'=ROUND(AVERAGE(F{idx}:K{idx}), 0)')
        ovr_cell.alignment = align_center
        ovr_cell.font = Font(name="Calibri", size=12, bold=True, color="000000")
        ovr_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

        # Tier: =IF(L{idx}>=75, "Ouro", IF(L{idx}>=65, "Prata", "Bronze"))
        tier_cell = ws3.cell(row=idx, column=13, value=f'=IF(L{idx}>=75, "Ouro", IF(L{idx}>=65, "Prata", "Bronze"))')
        tier_cell.alignment = align_center
        tier_cell.font = font_bold

        for c in range(1, 14):
            ws3.cell(row=idx, column=c).border = thin_border

    # Adjust column widths automatically
    for ws in [ws2, ws3]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 22
    ws3.column_dimensions['C'].width = 16
    ws3.column_dimensions['D'].width = 10
    ws3.column_dimensions['E'].width = 12

    wb.save(filename)
    print(f"Spreadsheet saved successfully at: {filename}")

if __name__ == "__main__":
    create_evaluation_spreadsheet(r"c:\Users\thiag\Desktop\plugshawtycafetoes FC\Planilha_Avaliacao_plugshawtycafetoes_FC.xlsx")
    create_evaluation_spreadsheet(r"C:\Users\thiag\Downloads\Planilha_Avaliacao_plugshawtycafetoes_FC.xlsx")
